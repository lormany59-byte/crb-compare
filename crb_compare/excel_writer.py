"""Step 6: build the template-styled comparison workbook."""

import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

from compare import summary_by_branch_currency

logger = logging.getLogger(__name__)

COLOR_HEADER_BG = "1E7A46"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_CUSTOMER = "C0561E"
COLOR_INCREASE = "1E7A46"
COLOR_DECREASE = "C0392B"

FMT_ACCOUNTING = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
FMT_TEXT = "@"

FONT_NAME = "Arial"


def _format_date(date_str: str) -> str:
    try:
        return f"{date_str[6:8]}/{date_str[4:6]}/{date_str[:4]}"
    except Exception:
        return date_str


def create_comparison_workbook(
    merged: pd.DataFrame,
    stats: dict,
    base_date: str,
    compare_date: str,
    base_lak: float,
    compare_lak: float,
    branch_map: dict,
    thresholds: dict,
    currency_order: list,
    output_raw_report: bool = False,
) -> Workbook:
    wb = Workbook()

    base_label = _format_date(base_date)
    compare_label = _format_date(compare_date)

    significant = merged[merged["is_significant"]]
    present_currencies = [c for c in currency_order if c in significant["CURRENCY"].unique()]

    _write_summary_sheet(
        wb, merged, stats, base_lak, compare_lak, currency_order, base_label, compare_label
    )

    for currency in present_currencies:
        cur_rows = significant[significant["CURRENCY"] == currency]

        cur_increase = cur_rows[cur_rows["change_type"] == "ເພີ່ມ"].sort_values(
            "diff", ascending=False
        )
        cur_decrease = cur_rows[cur_rows["change_type"] == "ຫຼຸດ"].sort_values(
            "diff", ascending=True
        )

        _write_change_sheet(
            wb, cur_increase, base_label, compare_label, f"ເງິນເພີ່ມຂື້ນ {currency}"
        )
        _write_change_sheet(
            wb, cur_decrease, base_label, compare_label, f"ເງິນທີ່ຫຼຸດລົງ {currency}"
        )

    if output_raw_report:
        _write_raw_report_sheet(wb, merged, base_label, compare_label)

    return wb


def write_excel(wb: Workbook, out_path) -> None:
    wb.save(out_path)
    logger.info(f"Output saved: {out_path}")


def _write_summary_sheet(
    wb, merged, stats, base_lak, compare_lak, currency_order, base_label, compare_label
):
    ws = wb.active
    ws.title = "ສະຫຼຸບ"

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=11)
    normal_font = Font(name=FONT_NAME, size=10)
    bold_font = Font(name=FONT_NAME, bold=True, size=10)
    title_font = Font(name=FONT_NAME, bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 1
    ws.cell(row=row, column=1, value="ລາຍງານສົມທຽບເງິນຝາກ CRB").font = title_font
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1).alignment = center
    row += 1

    ws.cell(
        row=row, column=1, value=f"ວັນທີ່ Base: {base_label}  |  ວັນທີ່ Compare: {compare_label}"
    ).font = Font(name=FONT_NAME, size=11)
    ws.merge_cells(f"A{row}:F{row}")
    row += 2

    ws.cell(row=row, column=1, value="ສະຫຼຸບຍອດ LAK").font = Font(
        name=FONT_NAME, bold=True, size=11
    )
    row += 1

    for col, h in enumerate(["ລາຍການ", "ຍອດ LAK"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.alignment = header_font, header_fill, center
    row += 1

    diff_lak = compare_lak - base_lak
    for label_val, amount in [
        (f"ຍອດ Base ({base_label})", base_lak),
        (f"ຍອດ Compare ({compare_label})", compare_lak),
        ("ການປ່ຽນແປງສຸດທິ (LAK)", diff_lak),
    ]:
        ws.cell(row=row, column=1, value=label_val).font = normal_font
        c = ws.cell(row=row, column=2, value=amount)
        c.number_format = FMT_ACCOUNTING
        c.font = Font(
            name=FONT_NAME, size=10, color=COLOR_INCREASE if amount >= 0 else COLOR_DECREASE
        )
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="ສະຖິຕິ").font = Font(name=FONT_NAME, bold=True, size=11)
    row += 1

    for col, h in enumerate(["ລາຍການ", "ຈຳນວນ"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.alignment = header_font, header_fill, center
    row += 1

    for label_val, count in [
        ("ບັນຊີທັງໝົດ", stats["total"]),
        ("ເພີ່ມຂຶ້ນ", stats["increase"]),
        ("ຫຼຸດລົງ", stats["decrease"]),
        ("ເທົ່າ", stats["equal"]),
        ("ເປີດໃໝ່", stats["new"]),
        ("ປິດ", stats["closed"]),
        ("ລາຍເຄື່ອນໄຫວໃຫຍ່", stats["significant"]),
    ]:
        ws.cell(row=row, column=1, value=label_val).font = normal_font
        ws.cell(row=row, column=2, value=count).font = bold_font
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="ສະຫຼຸບ ຕາມສາຂາ × ສະກຸນເງິນ").font = Font(
        name=FONT_NAME, bold=True, size=11
    )
    row += 1

    summary_headers = [
        "ສາຂາ",
        "ສະກຸນເງິນ",
        f"ຍອດ {base_label}",
        f"ຍອດ {compare_label}",
        "ການປ່ຽນແປງ",
        "ຈຳນວນບັນຊີ",
    ]
    for col, h in enumerate(summary_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font, c.fill, c.alignment = header_font, header_fill, center_wrap

    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(summary_headers))}{row}"
    ws.freeze_panes = f"A{row + 1}"
    row += 1

    summary = summary_by_branch_currency(merged, currency_order)
    for _, r in summary.iterrows():
        ws.cell(row=row, column=1, value=r["BranchName"]).font = normal_font
        ws.cell(row=row, column=2, value=r["CURRENCY"]).font = normal_font
        for col, val in [(3, r["base_sum"]), (4, r["compare_sum"]), (5, r["diff"])]:
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = FMT_ACCOUNTING
            c.font = normal_font
        ws.cell(row=row, column=6, value=int(r["count"])).font = normal_font
        row += 1

    widths = {1: 30, 2: 12, 3: 20, 4: 20, 5: 20, 6: 15}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_change_sheet(wb, df: pd.DataFrame, base_label, compare_label, sheet_name):
    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=10)
    normal_font = Font(name=FONT_NAME, size=10)
    customer_font = Font(name=FONT_NAME, size=10, color=COLOR_CUSTOMER)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "ສາຂາ",
        "ສົມທຽບວັນທີ",
        "ຊື່ບັນຊີ",
        "ເລກບັນຊີ",
        "ປະເພດທຸລະກິດ",
        "ສະກຸນເງິນ",
        f"ຍອດ {base_label}",
        f"ຍອດ {compare_label}",
        "ເງິນທີ່ເຂົ້າມາ / ອອກ",
        "ເຫດຜົນ",
    ]
    date_label = f"{base_label} ທຽບໃສ່ {compare_label}"

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = header_font, header_fill, center_wrap

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    data_start_row = 2
    row = data_start_row
    for _, r in df.iterrows():
        ws.cell(row=row, column=1, value=r["BranchName"]).font = normal_font
        ws.cell(row=row, column=2, value=date_label).font = normal_font

        c3 = ws.cell(row=row, column=3, value=r["CUSTOMER"])
        c3.font = customer_font

        c4 = ws.cell(row=row, column=4, value=r["CONTRACT"])
        c4.number_format = FMT_TEXT
        c4.font = normal_font

        ws.cell(row=row, column=5, value="").font = normal_font
        ws.cell(row=row, column=6, value=r["CURRENCY"]).font = normal_font

        for col, val in [(7, r["base_bal"]), (8, r["compare_bal"]), (9, r["diff"])]:
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = FMT_ACCOUNTING
            c.font = normal_font

        ws.cell(row=row, column=10, value="").font = normal_font
        row += 1

    data_end_row = row - 1

    if data_end_row >= data_start_row:
        ws.cell(row=row, column=1, value="ລວມທັງໝົດ").font = Font(
            name=FONT_NAME, bold=True, size=10
        )
        for col in (7, 8, 9):
            col_letter = get_column_letter(col)
            c = ws.cell(
                row=row,
                column=col,
                value=f"=SUBTOTAL(109,{col_letter}{data_start_row}:{col_letter}{data_end_row})",
            )
            c.number_format = FMT_ACCOUNTING
            c.font = Font(name=FONT_NAME, bold=True, size=10)

        col9 = get_column_letter(9)
        rng = f"{col9}{data_start_row}:{col9}{data_end_row}"

        pos_rule = Rule(
            type="cellIs",
            operator="greaterThan",
            formula=["0"],
            dxf=DifferentialStyle(font=Font(color=COLOR_INCREASE)),
        )
        neg_rule = Rule(
            type="cellIs",
            operator="lessThan",
            formula=["0"],
            dxf=DifferentialStyle(font=Font(color=COLOR_DECREASE)),
        )
        ws.conditional_formatting.add(rng, pos_rule)
        ws.conditional_formatting.add(rng, neg_rule)

    widths = {1: 25, 2: 22, 3: 35, 4: 20, 5: 18, 6: 12, 7: 20, 8: 20, 9: 20, 10: 20}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    logger.info(f"Sheet '{sheet_name}': {max(0, data_end_row - data_start_row + 1)} rows")


def _write_raw_report_sheet(wb, merged: pd.DataFrame, base_label, compare_label):
    ws = wb.create_sheet(title="Report")

    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=10)
    normal_font = Font(name=FONT_NAME, size=10)
    customer_font = Font(name=FONT_NAME, size=10, color=COLOR_CUSTOMER)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "BRANCH",
        "CONTRACT",
        "CURRENCY",
        "CUSTOMER",
        f"ຍອດ {base_label}",
        f"ຍອດ {compare_label}",
        "ຜິດດ່ຽງ",
        "ສະຖານະ",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = header_font, header_fill, center_wrap

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    for row_idx, (_, r) in enumerate(merged.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=r["BranchName"]).font = normal_font

        c = ws.cell(row=row_idx, column=2, value=r["CONTRACT"])
        c.number_format = FMT_TEXT
        c.font = normal_font

        ws.cell(row=row_idx, column=3, value=r["CURRENCY"]).font = normal_font
        ws.cell(row=row_idx, column=4, value=r["CUSTOMER"]).font = customer_font

        for col, key in [(5, "base_bal"), (6, "compare_bal"), (7, "diff")]:
            c = ws.cell(row=row_idx, column=col, value=r[key])
            c.number_format = FMT_ACCOUNTING
            c.font = normal_font

        status = []
        if r["is_new"]:
            status.append("ເປີດໃໝ່")
        if r["is_closed"]:
            status.append("ປິດ")
        ws.cell(
            row=row_idx, column=8, value=", ".join(status) or r["change_type"]
        ).font = normal_font

    widths = {1: 25, 2: 20, 3: 12, 4: 35, 5: 20, 6: 20, 7: 20, 8: 15}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
